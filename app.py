"""
QBO ↔ TicketVault Reconciliation — backend.

Two-way reconciliation between, for each entity:

  1. QuickBooks "Profit and Loss Detail"  — the accounting book. Income → Sales carries
     one Invoice line per marketplace per day (Name = "Stubhub (C)", "Vivid Seats (C)", …).
     Cost of Goods Sold carries one Journal Entry per day (a daily aggregate — no
     marketplace breakdown). A "Foreign Exchange Conversion" income sub-account, if
     present, is excluded entirely (its lines are never read).
  2. TicketVault export (CSV/XLSX)         — the operational book. Repeating daily blocks:
     a day row ("-> 07/01/2026") followed by per-client rows (TicketsNow, Vivid Seats,
     StubHub, …). The Sales "Net" column (Sold − Cancelled) is the sales figure; the Cost
     "Net" column is the cost figure.

Batch: upload any number of QBO files and any number of TicketVault files. Each QBO is
paired to its TicketVault by entity (from filename / alias table), then period, then by
total-sales proximity. The pairing is shown on the Summary tab.

Reconciliation, flagging any difference over a tolerance ($0.01 default):

  * SALES — per entity × marketplace × day. QBO's per-marketplace Invoice amount equals
            TicketVault's Sales-Net for that marketplace/day. Divergences are flagged;
            a marketplace present in only one book (non-zero) is flagged too.
  * COST  — per entity × day. QBO's daily COGS journal equals TicketVault's total Cost-Net
            for the day. (QBO has no per-marketplace cost, so cost is a day-level check.)

Output is lean: a Summary tab plus Sales Discrepancies and Cost Discrepancies tabs that
list only the flagged rows. Neither book is treated as the source of truth — both values
are shown side by side.
"""

import io
import os
import re
import csv
import time
import uuid
import shutil
import tempfile
import datetime as dt
from collections import defaultdict

from flask import Flask, request, jsonify, send_file, send_from_directory, abort
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__, static_folder=None)
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STORE_DIR = os.path.join(tempfile.gettempdir(), "qbo_tv_recon_store")
os.makedirs(STORE_DIR, exist_ok=True)


# =========================================================================== #
# RECONCILIATION CONFIG — edit these to tune matching
# =========================================================================== #

# Amounts within this many dollars are treated as equal (sub-penny never flagged).
TOLERANCE = 0.01

# Force differently-spelled marketplace names to be treated as one. Key = variant as it
# appears in either book (normalized: "(C)"/"(CAD)" stripped, lower-cased); value = the
# canonical display name. Only needed when a marketplace is labeled differently in QBO vs
# TicketVault; the common names already match once "(C)" is stripped.
MARKETPLACE_ALIASES = {
    # "stub hub": "StubHub",
}

# TicketVault client rows that are NOT QBO Sales marketplaces (internal transfers,
# expiries, etc.). Their COST still counts toward the day's cost total (QBO's daily COGS
# includes them), but they are not flagged as "missing" on the SALES side. Normalized,
# case-insensitive.
NON_MARKETPLACE_CLIENTS = {
    "baseball transfers",
    "expired",
    "transfers - yankees",
    "transfers",
}

# Map a QBO entity name or a TicketVault filename token to a canonical entity, so the two
# files pair up. Key = variant (normalized); value = canonical entity name. Matching is
# exact → suffix-tolerant (LLC/Inc/…) → this table. Add an entry whenever the Summary tab
# shows a file as unpaired.
ENTITY_ALIASES = {
    # "ys levine tickets": "YS Levine LLC",
}

# The company roster shown in the TicketVault drop-down — names and order exactly as in the
# Season Ticket Buy-In Review app. Each TicketVault file is tagged with one of these; the
# app maps it to the QBO file whose entity resolves to the same company.
COMPANIES = ["Y&S", "Grossman", "Sternbuch", "Pollak", "Levine", "Levovitz",
             "Chase", "Asher", "Katz", "GK", "TL", "Waxler", "TTG", "YourTix"]

# Force a QBO entity (or a stray spelling) onto a specific roster company when the automatic
# token match is wrong or missing. Key = entity text (normalized); value = a COMPANIES name.
# e.g. a QBO titled "YS Levine LLC" already resolves to "Levine" automatically; add an entry
# here only for names the matcher can't place.
COMPANY_ALIASES = {
    # "some holdings llc": "TTG",
}

# Authoritative name mapping from Master_Mapping_List.xlsx (QBO Company -> Short Name). The
# bundled Master_Mapping_List.xlsx, if present next to app.py, is read at startup and
# overrides/extends this — so the list can be updated by replacing the file. These entries
# are the fallback baked in so the app works without the file.
COMPANY_MAP_RAW = {
    "Damona & Crew": "Damona", "The Ticket Guy LLC": "TTG", "Y&S Tickets": "Y&S",
    "YourTickets": "YourTix", "YS Asher Tickets": "Asher", "YS Chase Tickets": "Chase",
    "YS Katz Tickets": "Katz", "YS Levine Tickets": "Levine",
    "YS Levovitz Tickets": "Levovitz", "YS Needle Tickets": "Needle",
    "YS TL Tickets": "TL", "YSKG Tickets": "GK", "YSM Tickets": "Grossman",
    "YSP Tickets": "Pollak", "YSS Tickets": "Sternbuch", "YSW Tickets": "Waxler",
}
# TicketVault-side spellings -> Short Name (extra keys that also resolve to a company).
COMPANY_TV_RAW = {
    "The Ticket Guy": "TTG", "YS Tickets": "Y&S", "YSA": "Asher", "Jacks YS": "Chase",
    "YS Katz": "Katz", "Yoni Levine": "Levine", "Needle Tickets LLC": "Needle",
    "YS TL": "TL", "GK LLC": "GK", "Pollak Tickets": "Pollak", "YSW": "Waxler",
    "Damon and Crew": "Damona",
}


# =========================================================================== #
# Generic helpers  (reused from the ledger-reconciliation scaffold)
# =========================================================================== #

_NUM_FORMULA = re.compile(r"-?\d+(\.\d+)?")
_MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun",
           "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def _cleanup_old(max_age_seconds=12 * 3600):
    now = time.time()
    for name in os.listdir(STORE_DIR):
        path = os.path.join(STORE_DIR, name)
        try:
            if os.path.isdir(path) and now - os.path.getmtime(path) > max_age_seconds:
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            pass


def _amount(v):
    """Coerce a cell to float. Handles numbers, '=123.45' literals, '$1,234.56',
    '(123)' negatives, stray whitespace. Cell-ref formulas (=B6+C6) -> None."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s in ("", " "):
        return None
    if s.startswith("="):
        body = s[1:]
        return float(body) if _NUM_FORMULA.fullmatch(body) else None
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in ("", "-", "."):
        return None
    try:
        f = float(s)
        return -f if neg else f
    except ValueError:
        return None


def _norm_date(x):
    """Normalize any date-ish cell to a datetime.date."""
    if isinstance(x, dt.datetime):
        return x.date()
    if isinstance(x, dt.date):
        return x
    if isinstance(x, str):
        s = x.strip()
        m = re.search(r"(\d{1,2}/\d{1,2}/\d{2,4})", s)   # e.g. " -> 07/01/2026"
        if m:
            s = m.group(1)
        for f in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y", "%B %d, %Y", "%b %d, %Y"):
            try:
                return dt.datetime.strptime(s, f).date()
            except ValueError:
                continue
    return None


def _norm_name(s):
    return re.sub(r"\s+", " ", str(s).strip()).lower() if s is not None else ""


def _norm_marketplace(s):
    """Display name for a marketplace cell, with the QBO customer/CAD markers removed."""
    if s is None:
        return "(unlabeled)"
    t = str(s).strip()
    t = re.sub(r"\((?:C|c|CAD|cad)\)", "", t)
    t = re.sub(r"\s+", " ", t).strip()
    return t if t else "(unlabeled)"


def _mkt_key(raw):
    """(casefold key, display) for a marketplace — case-insensitive grouping with alias."""
    disp = _norm_marketplace(raw)
    key = disp.casefold()
    if key in MARKETPLACE_ALIASES:
        disp = MARKETPLACE_ALIASES[key]
        key = disp.casefold()
    return key, disp


def _rows_from_upload(filename, data):
    """Return list-of-lists of cell values from a .csv/.xlsx/.xlsm upload,
    picking the most-populated worksheet."""
    low = filename.lower()
    if low.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    best, best_score = None, -1
    for ws in wb.worksheets:
        score = (ws.max_row or 0) * (ws.max_column or 1)
        if score > best_score:
            best, best_score = ws, score
    rows = [list(r) for r in best.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_with_indent(filename, data):
    """Like _rows_from_upload but bakes each row's indentation (literal leading spaces +
    Excel indent property) into the first text cell, so the QBO account tree can be scoped
    by depth. CSV falls back to plain rows (QBO exports keep literal leading spaces)."""
    if filename.lower().endswith(".csv"):
        return _rows_from_upload(filename, data)
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    best, best_score = None, -1
    for ws in wb.worksheets:
        score = (ws.max_row or 0) * (ws.max_column or 1)
        if score > best_score:
            best, best_score = ws, score
    out = []
    for row in best.iter_rows():
        first_text = next((c for c in row if isinstance(c.value, str) and c.value.strip()),
                          None)
        vals = []
        for c in row:
            v = c.value
            if c is first_text:
                raw = str(v)
                lead = len(raw) - len(raw.lstrip())
                try:
                    align = int(round(c.alignment.indent or 0))
                except (TypeError, ValueError):
                    align = 0
                v = (" " * (lead + align * 3)) + raw.strip()
            vals.append(v)
        out.append(vals)
    wb.close()
    return out


def _find_header(rows, must_have, limit=15):
    """Return (index, header_cells_lowered_list) for the first row within `limit` rows
    that contains all of `must_have` (lowered substrings)."""
    want = [w.lower() for w in must_have]
    for i, row in enumerate(rows[:limit]):
        cells = [(_norm_name(c) if c is not None else "") for c in row]
        if all(any(w == c or w in c for c in cells) for w in want):
            return i, cells
    return None, []


def _report_period(rows):
    """Find a 'Month D-D, YYYY' style period line in the top rows, else None."""
    for row in rows[:6]:
        for cell in row[:3]:
            if isinstance(cell, str) and re.search(r"\b20\d\d\b", cell) and \
                    any(m in cell for m in _MONTHS):
                return cell.strip()
    return None


def _period_end(rows):
    """Best-effort month-end date from a period line; else None."""
    line = _report_period(rows)
    if line:
        m = re.search(r"([A-Za-z]+)\s+\d+\s*-\s*(\d+),\s*(\d{4})", line)
        if m:
            try:
                return dt.datetime.strptime(f"{m.group(1)} {m.group(2)}, {m.group(3)}",
                                            "%B %d, %Y").date()
            except ValueError:
                pass
    return None


_REPORT_TITLES = {"transaction report", "general ledger", "balance sheet",
                  "profit and loss detail", "profit and loss", "profit & loss detail"}


def _detect_entity(rows):
    """First real text cell in the top rows that isn't a report title, an 'As of' line,
    or a date/period line — i.e. the company/entity the report was run for."""
    for row in rows[:5]:
        for cell in row:
            if isinstance(cell, str) and cell.strip():
                s = cell.strip()
                low = s.lower()
                if low in _REPORT_TITLES or low.startswith("as of"):
                    break
                if _report_period([[s]]):
                    break
                return s
            elif cell is not None and str(cell).strip():
                break
    return None


_SUFFIX_RE = re.compile(r"\b(l\.?l\.?c\.?|inc\.?|incorporated|corp\.?|co\.?|company|ltd\.?)\b")


def _strip_suffix(name):
    return re.sub(r"\s+", " ", _SUFFIX_RE.sub("", _norm_name(name))).strip()


def _canon_entity(name):
    """Canonicalize an entity name via the alias table (else keep as given)."""
    if not name:
        return None
    key = _norm_name(name)
    if key in ENTITY_ALIASES:
        return ENTITY_ALIASES[key]
    skey = _strip_suffix(name)
    for k, v in ENTITY_ALIASES.items():
        if _strip_suffix(k) == skey and skey:
            return v
    return str(name).strip()


def _entities_match(a, b):
    """True if two entity names refer to the same entity (exact / suffix-tolerant / alias)."""
    if not a or not b:
        return False
    ca, cb = _canon_entity(a), _canon_entity(b)
    if _norm_name(ca) == _norm_name(cb):
        return True
    return bool(_strip_suffix(ca)) and _strip_suffix(ca) == _strip_suffix(cb)


def _company_core(label):
    """Normalized single-token core of a roster company ('Y&S' -> 'ys', 'YourTix' ->
    'yourtix')."""
    return re.sub(r"[^a-z0-9]", "", str(label).lower())


# Words dropped when reducing a company name to its distinctive "core", so that a QBO title
# ("YS Levine LLC") and a mapping entry ("YS Levine Tickets") reduce to the same key.
_MAP_STOPWORDS = {"llc", "inc", "incorporated", "corp", "co", "company", "ltd", "the",
                  "tickets", "ticket", "tix"}


def _map_core(name):
    norm = _norm_name(name).replace("&", "").replace("+", "")
    toks = [t for t in re.findall(r"[a-z0-9]+", norm) if t not in _MAP_STOPWORDS]
    return "".join(toks)


def _load_mapping_file():
    """Read Master_Mapping_List.xlsx if it's bundled next to app.py, returning
    {core: short_name}. Reads the QBO-Company column and the TicketVault-Company column,
    both mapped to the Short-Name column. Returns {} if the file isn't present."""
    path = os.path.join(BASE_DIR, "Master_Mapping_List.xlsx")
    out = {}
    if not os.path.exists(path):
        return out
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        for r in wb.worksheets[0].iter_rows(values_only=True):
            r = list(r)
            short = str(r[1]).strip() if len(r) > 1 and r[1] and str(r[1]).strip() else None
            if not short:
                continue
            for col in (0, 3):    # QBO Company, TicketVault Company
                v = r[col] if len(r) > col else None
                if v and str(v).strip().upper() != "N/A":
                    core = _map_core(v)
                    if core:
                        out[core] = short
        wb.close()
    except Exception:
        pass
    return out


# core -> short name, built from the baked-in tables then overridden by the bundled file
MAP_BY_CORE = {}
for _raw in (COMPANY_MAP_RAW, COMPANY_TV_RAW):
    for _k, _v in _raw.items():
        _c = _map_core(_k)
        if _c:
            MAP_BY_CORE[_c] = _v
MAP_BY_CORE.update(_load_mapping_file())
MAP_BY_CORE.pop("", None)


def resolve_company(name):
    """Map an entity name (or filename stem, or a picked value) to one of the COMPANIES
    roster names, using the master mapping list first. Order: exact roster name →
    COMPANY_ALIASES → master-mapping core → longest roster token in the name. None if
    nothing fits."""
    if not name:
        return None
    norm = _norm_name(name)
    for c in COMPANIES:                     # already a roster name?
        if _norm_name(c) == norm:
            return c
    if norm in COMPANY_ALIASES:             # explicit override
        return COMPANY_ALIASES[norm]
    skey = _strip_suffix(name)
    for k, v in COMPANY_ALIASES.items():
        if _strip_suffix(k) == skey and skey:
            return v
    core = _map_core(name)                  # master mapping list (authoritative)
    if core in MAP_BY_CORE:
        return MAP_BY_CORE[core]
    # token fallback — a roster core equal to one of the name's tokens; longest core wins.
    joined = norm.replace("&", "").replace("+", "")
    tokens = set(re.findall(r"[a-z0-9]+", joined))
    best, best_len = None, 0
    for c in COMPANIES:
        cc = _company_core(c)
        if cc in tokens and len(cc) > best_len:
            best, best_len = c, len(cc)
    return best


# =========================================================================== #
# Parsers
# =========================================================================== #

def _looks_like_qbo(rows):
    """Heuristic: does this look like a QuickBooks P&L/Detail rather than a TicketVault
    export? (Has a 'Profit and Loss' title, or a Date/Amount/Balance header.)"""
    for row in rows[:8]:
        for c in row:
            if isinstance(c, str) and "profit and loss" in c.lower():
                return True
    hi, _ = _find_header(rows, ["amount", "balance"])
    return hi is not None


def _looks_like_ticketvault(rows):
    """Heuristic: does this look like a TicketVault export? (Has a Client + Sold header.)"""
    for row in rows[:12]:
        cells = [(_norm_name(c) if c is not None else "") for c in row]
        if "client" in cells and "sold" in cells:
            return True
    return False


def parse_qbo(filename, data):
    """QuickBooks 'Profit and Loss Detail' -> per-entity book.

    Returns (entity, book, meta) where book = {
        "sales": {(mkt_key, date): amount},   # Income → Sales invoice lines
        "cost":  {date: amount},              # Cost of Goods Sold daily journals
        "mkt_names": {mkt_key: display},
        "sales_total": float, "cost_total": float,
    }
    Foreign Exchange Conversion lines are excluded entirely.
    """
    rows = _rows_with_indent(filename, data)
    hi, _ = _find_header(rows, ["date", "amount", "balance"])
    if hi is None:
        if _looks_like_ticketvault(rows):
            raise ValueError(f"'{filename}' looks like a TicketVault export, not a "
                             f"QuickBooks P&L Detail — put it in the TicketVault box.")
        raise ValueError(f"QBO '{filename}': could not find a header row with "
                         f"'Date', 'Amount' and 'Balance'.")
    hdr = [(_norm_name(c) if c is not None else "") for c in rows[hi]]

    def col(*names):
        for n in names:
            for j, c in enumerate(hdr):
                if c == n or (n and n in c):
                    return j
        return None

    c_date = col("date")
    c_name = col("name")
    c_amt = col("amount")

    book = {"sales": defaultdict(float), "cost": defaultdict(float), "mkt_names": {}}
    section = None            # "sales" | "cost" | None

    for row in rows[hi + 1:]:
        a = row[0] if row else None
        if isinstance(a, str) and a.strip():
            lab = a.strip()
            depth = (len(a) - len(a.lstrip())) // 3
            low = lab.lower()
            if low.startswith("total for"):
                section = None
            elif low == "sales":
                section = "sales"
            elif low == "cost of goods sold":
                # the COGS account holds one journal entry per day. QBO exports vary in
                # whether the account tree is indented, so match by name, not by depth.
                section = "cost"
            else:
                # any other node / parent header — including "Foreign Exchange
                # Conversion", which is deliberately excluded from reconciliation.
                section = None
            continue
        # ---- data row (col A empty) ----
        if section is None:
            continue
        amt = _amount(row[c_amt]) if c_amt is not None and c_amt < len(row) else None
        if amt is None:
            continue
        d = _norm_date(row[c_date]) if c_date is not None and c_date < len(row) else None
        if d is None:
            continue
        if section == "sales":
            raw = row[c_name] if c_name is not None and c_name < len(row) else None
            mkey, mdisp = _mkt_key(raw)
            book["sales"][(mkey, d)] += amt
            book["mkt_names"].setdefault(mkey, mdisp)
        elif section == "cost":
            book["cost"][d] += amt

    entity = _canon_entity(_detect_entity(rows))
    book["sales_total"] = round(sum(book["sales"].values()), 2)
    book["cost_total"] = round(sum(book["cost"].values()), 2)
    book["sales"] = dict(book["sales"])
    book["cost"] = dict(book["cost"])
    meta = {"period": _report_period(rows), "period_end": _period_end(rows),
            "entity": entity, "filename": filename}
    return entity, book, meta


_TV_HEADER_WORDS = {"sales", "cost", "invoices", "tickets", "date", "client", "net",
                    "sold", "cancelled", "cost of sold", "cost of cancelled",
                    "net profit", "profit"}


def _tv_entity(rows, hi):
    """Look for a company/title row above the TicketVault header. A real title row has a
    single populated cell that isn't a column-group word or a date/period line. Most
    TicketVault exports have no such row (returns None), in which case pairing falls back
    to the filename / period / totals."""
    for row in rows[:hi]:
        populated = [str(c).strip() for c in row if c is not None and str(c).strip()]
        if len(populated) != 1:
            continue
        cand = populated[0]
        low = cand.lower()
        if low in _TV_HEADER_WORDS or not re.search(r"[A-Za-z]", cand):
            continue
        if _report_period([[cand]]) or _norm_date(cand):
            continue
        return cand
    return None


def parse_ticketvault(filename, data):
    """TicketVault export -> book of the same shape as parse_qbo's (minus fx).

    Day blocks: a row with text in column 0 ('-> MM/DD/YYYY') is the day header/total;
    the following rows (blank column 0, client in column 1) are the per-client lines.
    Sales uses the Sales-group 'Net' column (Sold − Cancelled); cost uses the Cost-group
    'Net' column. Client cost is summed per day to match QBO's daily COGS aggregate.
    """
    rows = _rows_from_upload(filename, data)
    # Header row: the one containing 'client' + 'sold' (+ 'cost of sold').
    hi = None
    for i, row in enumerate(rows[:12]):
        cells = [(_norm_name(c) if c is not None else "") for c in row]
        if "client" in cells and "sold" in cells:
            hi = i
            hdr = cells
            break
    if hi is None:
        if _looks_like_qbo(rows):
            raise ValueError(f"'{filename}' looks like a QuickBooks P&L Detail, not a "
                             f"TicketVault export. Put your TicketVault export in this box "
                             f"(and P&L files go in the QuickBooks box above).")
        raise ValueError(f"TicketVault '{filename}': could not find the "
                         f"'Date / Client / Sold …' header row.")

    def idx(name):
        return hdr.index(name) if name in hdr else None

    c_date = idx("date") if idx("date") is not None else 0
    c_client = idx("client") if idx("client") is not None else 1
    # Two 'net' columns: sales-net after 'cancelled', cost-net after 'cost of cancelled'.
    net_positions = [j for j, c in enumerate(hdr) if c == "net"]
    i_cancelled = hdr.index("cancelled") if "cancelled" in hdr else 3
    i_cost_can = hdr.index("cost of cancelled") if "cost of cancelled" in hdr else 6
    c_sales_net = next((j for j in net_positions if j > i_cancelled), 4)
    c_cost_net = next((j for j in net_positions if j > i_cost_can), 7)

    entity = _canon_entity(_tv_entity(rows, hi))
    book = {"sales": defaultdict(float), "cost": defaultdict(float), "mkt_names": {}}
    cur_date = None
    for row in rows[hi + 1:]:
        first = row[c_date] if c_date < len(row) else None
        if first is not None and str(first).strip():
            cur_date = _norm_date(first)          # day header/total row — sets the date
            continue
        client = row[c_client] if c_client < len(row) else None
        if client is None or not str(client).strip() or cur_date is None:
            continue
        sales_net = _amount(row[c_sales_net]) if c_sales_net < len(row) else None
        cost_net = _amount(row[c_cost_net]) if c_cost_net < len(row) else None
        mkey, mdisp = _mkt_key(client)
        if cost_net is not None:
            book["cost"][cur_date] += cost_net
        if _norm_name(client) in NON_MARKETPLACE_CLIENTS:
            continue                              # not a QBO sales marketplace
        if sales_net is not None:
            book["sales"][(mkey, cur_date)] += sales_net
            book["mkt_names"].setdefault(mkey, mdisp)

    all_dates = [d for (_, d) in book["sales"]] + list(book["cost"])
    meta = {"period_end": max(all_dates) if all_dates else None,
            "period_start": min(all_dates) if all_dates else None,
            "period": None, "entity": entity, "filename": filename}
    if all_dates:
        s, e = min(all_dates), max(all_dates)
        meta["period"] = (s.strftime("%B %-d") + "-" + e.strftime("%-d, %Y")
                          if s.month == e.month and s.year == e.year
                          else f"{s:%m/%d/%Y}-{e:%m/%d/%Y}")
    book["sales_total"] = round(sum(book["sales"].values()), 2)
    book["cost_total"] = round(sum(book["cost"].values()), 2)
    book["sales"] = dict(book["sales"])
    book["cost"] = dict(book["cost"])
    return book, meta


# =========================================================================== #
# Pairing  (QBO file  <->  TicketVault file)
# =========================================================================== #

def _periods_overlap(a, b):
    (a0, a1), (b0, b1) = a, b
    if None in (a0, a1, b0, b1):
        return False
    return a0 <= b1 and b0 <= a1


def _qbo_period_range(book, meta):
    dates = [d for (_, d) in book["sales"]] + list(book["cost"])
    if not dates:
        return (None, None)
    return (min(dates), max(dates))


def pair_files(qbos, tvs):
    """qbos: list of (entity, book, meta). tvs: list of (book, meta, entity_from_name).
    Greedy one-to-one pairing by entity match, then period overlap, then total-sales
    proximity. Returns (pairs, unpaired_qbo, unpaired_tv, basis) where pairs is a list of
    (qi, ti) index tuples and basis[(qi,ti)] describes why."""
    cand = []
    for qi, (qent, qbook, qmeta) in enumerate(qbos):
        qrange = _qbo_period_range(qbook, qmeta)
        for ti, (tbook, tmeta, tent) in enumerate(tvs):
            trange = (tmeta.get("period_start"), tmeta.get("period_end"))
            score, why = 0, []
            if qent and tent and _entities_match(qent, tent):
                score += 100
                why.append("entity")
            if qrange == trange and None not in qrange:
                score += 50
                why.append("period=")
            elif _periods_overlap(qrange, trange):
                score += 25
                why.append("period~")
            sdiff = abs((qbook["sales_total"] or 0) - (tbook["sales_total"] or 0))
            if sdiff <= TOLERANCE:
                score += 40
                why.append("sales=")
            else:
                score += max(0, 20 - min(20, sdiff / 1000.0))
            cand.append((score, qi, ti, ", ".join(why) or "best-fit"))

    cand.sort(reverse=True)
    used_q, used_t, pairs, basis = set(), set(), [], {}
    for score, qi, ti, why in cand:
        if qi in used_q or ti in used_t:
            continue
        used_q.add(qi)
        used_t.add(ti)
        pairs.append((qi, ti))
        basis[(qi, ti)] = why
    unpaired_q = [i for i in range(len(qbos)) if i not in used_q]
    unpaired_t = [i for i in range(len(tvs)) if i not in used_t]
    return pairs, unpaired_q, unpaired_t, basis


# =========================================================================== #
# Reconciliation
# =========================================================================== #

def _eq(a, b):
    return abs((a or 0.0) - (b or 0.0)) <= TOLERANCE


def reconcile_pair(entity, qbook, tbook):
    """Return (sales_rows, cost_rows) of flagged discrepancies for one entity."""
    sales_rows, cost_rows = [], []

    # ---- Sales: marketplace × day ----
    keys = set(qbook["sales"]) | set(tbook["sales"])
    names = {}
    names.update(tbook.get("mkt_names", {}))
    names.update(qbook.get("mkt_names", {}))
    for (mkey, d) in sorted(keys, key=lambda k: (k[1], names.get(k[0], k[0]).lower())):
        q = qbook["sales"].get((mkey, d))
        v = tbook["sales"].get((mkey, d))
        if _eq(q or 0.0, v or 0.0):
            continue
        if q is None or abs(q) <= TOLERANCE:
            status = "TicketVault only"
        elif v is None or abs(v) <= TOLERANCE:
            status = "QBO only"
        else:
            status = "Amount Mismatch"
        sales_rows.append({
            "Entity": entity, "Date": d, "Marketplace": names.get(mkey, mkey),
            "QBO": round(q, 2) if q is not None else 0.0,
            "TicketVault": round(v, 2) if v is not None else 0.0,
            "Diff (QBO-TV)": round((q or 0.0) - (v or 0.0), 2),
            "Status": status, "_flag": True,
        })

    # ---- Cost: day level ----
    for d in sorted(set(qbook["cost"]) | set(tbook["cost"])):
        q = qbook["cost"].get(d)
        v = tbook["cost"].get(d)
        if _eq(q or 0.0, v or 0.0):
            continue
        if q is None or abs(q) <= TOLERANCE:
            status = "TicketVault only"
        elif v is None or abs(v) <= TOLERANCE:
            status = "QBO only"
        else:
            status = "Amount Mismatch"
        cost_rows.append({
            "Entity": entity, "Date": d,
            "QBO Cost": round(q, 2) if q is not None else 0.0,
            "TicketVault Cost": round(v, 2) if v is not None else 0.0,
            "Diff (QBO-TV)": round((q or 0.0) - (v or 0.0), 2),
            "Status": status, "_flag": True,
        })
    return sales_rows, cost_rows


# =========================================================================== #
# Workbook builder
# =========================================================================== #

HEAD_FILL = PatternFill("solid", fgColor="374151")
HEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=13)
SUB_FONT = Font(name="Arial", italic=True, color="6B7280", size=9)
LABEL_FONT = Font(name="Arial", bold=True, size=10)
BASE_FONT = Font(name="Arial", size=10)
MONEY = '#,##0.00;(#,##0.00);"-"'
THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(bottom=THIN)


def _style_header(ws, row_idx, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _emit_table(ws, start_row, columns, rows, money_cols, total_cols=None):
    """Write a header + data table. Returns the row after the table (incl. total)."""
    ncols = len(columns)
    for j, name in enumerate(columns, start=1):
        ws.cell(row=start_row, column=j, value=name)
    _style_header(ws, start_row, ncols)
    r = start_row + 1
    first_data = r
    for rec in rows:
        for j, name in enumerate(columns, start=1):
            v = rec.get(name, "")
            if isinstance(v, dt.date):
                v = v.strftime("%m/%d/%Y")
            cell = ws.cell(row=r, column=j, value=(v if v is not None else ""))
            cell.font = BASE_FONT
            cell.border = BORDER
            if name in money_cols and isinstance(rec.get(name), (int, float)):
                cell.number_format = MONEY
            if name in money_cols:
                cell.alignment = Alignment(horizontal="right")
        r += 1
    last_data = r - 1
    if total_cols and last_data >= first_data:
        ws.cell(row=r, column=1, value="TOTAL").font = LABEL_FONT
        for name in total_cols:
            j = columns.index(name) + 1
            col = get_column_letter(j)
            cell = ws.cell(row=r, column=j,
                           value=f"=SUM({col}{first_data}:{col}{last_data})")
            cell.font = LABEL_FONT
            cell.number_format = MONEY
            cell.alignment = Alignment(horizontal="right")
        r += 1
    return r


def _qbo_label(entity, filename):
    """The label shown for a QBO file in the company dropdown: its detected entity, or the
    filename stem when the entity can't be read."""
    return entity or os.path.splitext(os.path.basename(filename))[0]


def detect_qbo_entity(filename, data):
    """Light read of just the entity name from a QBO file's title row (for /qbo_entities).
    Returns the label the dropdown should show."""
    try:
        rows = _rows_from_upload(filename, data)
    except Exception:
        return _qbo_label(None, filename)
    return _qbo_label(_canon_entity(_detect_entity(rows)), filename)


def build_workbook(qbo_files, tv_files, tv_companies=None):
    """qbo_files / tv_files: lists of (filename, bytes). tv_companies: optional list, one
    entry per TicketVault file, of the entity the user picked in the dropdown (authoritative
    when present)."""
    qbos, tvs, warnings = [], [], []

    for fn, data in qbo_files:
        entity, book, meta = parse_qbo(fn, data)
        if not entity:
            warnings.append(f"QBO '{fn}': could not detect the entity name from its "
                            f"title row — pairing will rely on period / totals.")
        qbos.append((entity, book, meta))

    qlabels = [_qbo_label(q[0], q[2]["filename"]) for q in qbos]
    qcompanies = [resolve_company(q[0]) or resolve_company(q[2]["filename"]) for q in qbos]
    known_entities = [q[0] for q in qbos if q[0]]
    tv_companies = tv_companies or []
    for i, (fn, data) in enumerate(tv_files):
        book, meta = parse_ticketvault(fn, data)
        selected = (tv_companies[i] if i < len(tv_companies) else "") or ""
        selected = selected.strip()
        if selected:
            # dropdown pick — a roster company, authoritative.
            meta["company"] = resolve_company(selected) or selected
            meta["entity"] = meta["company"]
            meta["selected"] = True
        else:
            # 1) company name read from inside the file, if any; else 2) a filename token.
            ent = meta.get("entity")
            if not ent:
                stem = os.path.splitext(os.path.basename(fn))[0].replace("_", " ")
                for kent in known_entities:
                    if _entities_match(stem, kent):
                        ent = kent
                        break
            meta["entity"] = ent
            meta["company"] = resolve_company(ent) or resolve_company(fn)
            meta["selected"] = False
        tvs.append((book, meta, meta["entity"]))

    # ---- pairing: honor explicit dropdown selections first, auto-pair the rest ----
    used_q, used_t, pairs, basis = set(), set(), [], {}
    for ti, (tbook, tmeta, tent) in enumerate(tvs):
        if not tmeta.get("selected"):
            continue
        sel = tmeta["entity"]                       # a roster company name
        sel_co = tmeta.get("company") or sel
        # match to the QBO whose entity resolves to the same roster company (then fall back
        # to a direct name / suffix match).
        match_qi = next(
            (qi for qi in range(len(qbos))
             if qi not in used_q and
             ((qcompanies[qi] and _norm_name(qcompanies[qi]) == _norm_name(sel_co))
              or _norm_name(qlabels[qi]) == _norm_name(sel)
              or _entities_match(qlabels[qi], sel))), None)
        if match_qi is None:
            warnings.append(f"TicketVault '{tmeta['filename']}' is assigned to '{sel}', "
                            f"which doesn't match any uploaded QBO file — not reconciled.")
            used_t.add(ti)
            continue
        used_q.add(match_qi)
        used_t.add(ti)
        pairs.append((match_qi, ti))
        basis[(match_qi, ti)] = "company (selected)"

    rem_q = [qi for qi in range(len(qbos)) if qi not in used_q]
    rem_t = [ti for ti in range(len(tvs)) if ti not in used_t]
    gpairs, gup_q, gup_t, gbasis = pair_files([qbos[i] for i in rem_q],
                                              [tvs[i] for i in rem_t])
    for (lqi, lti) in gpairs:
        gqi, gti = rem_q[lqi], rem_t[lti]
        pairs.append((gqi, gti))
        basis[(gqi, gti)] = gbasis[(lqi, lti)]
    unpaired_q = [rem_q[i] for i in gup_q]
    unpaired_t = [rem_t[i] for i in gup_t]

    sales_rows, cost_rows, pair_info = [], [], []
    for qi, ti in pairs:
        qent, qbook, qmeta = qbos[qi]
        tbook, tmeta, _ = tvs[ti]
        entity = (qcompanies[qi] or tmeta.get("company")
                  or qent or tmeta.get("entity") or f"(entity {qi + 1})")
        s, c = reconcile_pair(entity, qbook, tbook)
        sales_rows.extend(s)
        cost_rows.extend(c)
        pair_basis = basis.get((qi, ti), "")
        pair_info.append({
            "Entity": entity,
            "QBO File": qmeta["filename"],
            "TicketVault File": tmeta["filename"],
            "Period": qmeta.get("period") or tmeta.get("period") or "",
            "Sales Δ": len(s), "Cost Δ": len(c),
            "Matched On": pair_basis,
            "_flag": bool(s or c),
        })
        # In a batch, a pair matched only by period/amount (no confirmed company name and
        # no dropdown selection) is a guess — surface it so the user can verify.
        if (len(qbos) > 1 or len(tvs) > 1) and "entity" not in pair_basis \
                and "selected" not in pair_basis:
            warnings.append(
                f"TicketVault '{tmeta['filename']}' was paired to '{entity}' by "
                f"period/amount only — no company name was found in the file or its "
                f"filename. Verify this pairing, or name the file with the company (or "
                f"add an ENTITY_ALIASES entry).")
    for qi in unpaired_q:
        qent, _, qmeta = qbos[qi]
        warnings.append(f"QBO '{qmeta['filename']}' (entity '{qent or 'unknown'}') had no "
                        f"matching TicketVault export — not reconciled.")
    for ti in unpaired_t:
        _, tmeta, _ = tvs[ti]
        warnings.append(f"TicketVault '{tmeta['filename']}' had no matching QBO file — "
                        f"not reconciled.")

    # sort discrepancy rows for readability
    sales_rows.sort(key=lambda x: (x["Entity"].lower(), x["Date"], x["Marketplace"].lower()))
    cost_rows.sort(key=lambda x: (x["Entity"].lower(), x["Date"]))

    period_end = next((q[2].get("period_end") for q in qbos if q[2].get("period_end")),
                      None) or dt.date.today()

    # ------------------------------------------------------------------ workbook
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Sales Discrepancies ----
    ws = wb.create_sheet("Sales Discrepancies")
    scols = ["Entity", "Date", "Marketplace", "QBO", "TicketVault", "Diff (QBO-TV)", "Status"]
    smoney = {"QBO", "TicketVault", "Diff (QBO-TV)"}
    if sales_rows:
        _emit_table(ws, 1, scols, sales_rows, smoney, total_cols=["QBO", "TicketVault", "Diff (QBO-TV)"])
        ws.freeze_panes = "A2"
    else:
        ws["A1"] = "No sales discrepancies — every marketplace/day reconciled within tolerance."
        ws["A1"].font = BASE_FONT
    _autofit(ws, [22, 12, 20, 14, 14, 15, 18])

    # ---- Cost Discrepancies ----
    ws = wb.create_sheet("Cost Discrepancies")
    ccols = ["Entity", "Date", "QBO Cost", "TicketVault Cost", "Diff (QBO-TV)", "Status"]
    cmoney = {"QBO Cost", "TicketVault Cost", "Diff (QBO-TV)"}
    if cost_rows:
        _emit_table(ws, 1, ccols, cost_rows, cmoney,
                    total_cols=["QBO Cost", "TicketVault Cost", "Diff (QBO-TV)"])
        ws.freeze_panes = "A2"
    else:
        ws["A1"] = "No cost discrepancies — every day reconciled within tolerance."
        ws["A1"].font = BASE_FONT
    _autofit(ws, [22, 12, 16, 18, 15, 18])

    buf = io.BytesIO()
    wb.save(buf)

    meta = {
        "period_month": period_end.strftime("%B %Y"),
        "entities": len(pairs),
        "sales_flags": len(sales_rows),
        "cost_flags": len(cost_rows),
        "unpaired": len(unpaired_q) + len(unpaired_t),
        "warnings": warnings,
    }
    return buf.getvalue(), meta


# =========================================================================== #
# Routes
# =========================================================================== #

@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/options")
def options():
    """The company roster for the TicketVault drop-down — names and order from config."""
    return jsonify({"companies": COMPANIES})


@app.route("/process", methods=["POST"])
def process():
    qbo_files = [(f.filename, f.read())
                 for f in request.files.getlist("qbo") if f.filename]
    tv_files = [(f.filename, f.read())
                for f in request.files.getlist("ticketvault") if f.filename]
    tv_companies = request.form.getlist("ticketvault_company")
    if not qbo_files:
        return jsonify({"error": "Please upload at least one QuickBooks P&L Detail."}), 400
    if not tv_files:
        return jsonify({"error": "Please upload at least one TicketVault export."}), 400

    try:
        data, meta = build_workbook(qbo_files, tv_files, tv_companies)
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    except Exception as exc:
        return jsonify({"error": f"{type(exc).__name__}: {exc}"}), 500

    token = uuid.uuid4().hex
    folder = os.path.join(STORE_DIR, token)
    os.makedirs(folder, exist_ok=True)
    fn = f"QBO vs TicketVault Recon {meta['period_month']}.xlsx"
    with open(os.path.join(folder, fn), "wb") as fh:
        fh.write(data)
    _cleanup_old()

    warnings = []
    if meta["sales_flags"]:
        warnings.append(f"{meta['sales_flags']} sales discrepancy row(s) flagged.")
    if meta["cost_flags"]:
        warnings.append(f"{meta['cost_flags']} cost discrepancy row(s) flagged.")
    if not meta["sales_flags"] and not meta["cost_flags"] and meta["entities"]:
        warnings.append("Everything reconciled within tolerance — no discrepancies.")
    warnings.extend(meta["warnings"])

    return jsonify({
        "download_url": f"/download/{token}",
        "filename": fn,
        "entities": meta["entities"],
        "warnings": warnings,
    })


@app.route("/download/<token>")
def download(token):
    folder = os.path.join(STORE_DIR, os.path.basename(token))
    if not os.path.isdir(folder):
        abort(404)
    xlsx = [f for f in os.listdir(folder) if f.lower().endswith(".xlsx")]
    if not xlsx:
        abort(404)
    pick = xlsx[0]
    return send_file(os.path.join(folder, pick),
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=pick)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
